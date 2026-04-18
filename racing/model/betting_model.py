"""
HORSE RACING PREDICTION MODEL v2
==================================
Upload your Daily.xlsx each morning and run this script.
Trains on 9 years of historical results, scores today's runners,
outputs ranked back and lay bets with expected profit.

USAGE:
    python betting_model.py Daily.xlsx

SETUP:
    python -m pip install pandas openpyxl scikit-learn numpy
"""

import sys
import warnings
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import cross_val_score, StratifiedKFold
warnings.filterwarnings("ignore")

# ── CONFIG ────────────────────────────────────────────────────────────────────
MIN_BACK_ODDS  = 2.5    # Only back at these odds or above
MAX_BACK_ODDS  = 15.0   # Don't back massive outsiders
MIN_LAY_ODDS   = 2.5    # Only lay at these odds or above
MAX_LAY_ODDS   = 20.0   # Don't expose yourself to huge liabilities
TOP_N_BETS     = 8      # Max bets to show per category
MIN_EDGE       = 0.04   # Minimum 4% edge over market to recommend
BETFAIR_COMM   = 0.02   # Betfair 2% commission

# ── LOAD ──────────────────────────────────────────────────────────────────────

def load_excel(filepath):
    print("Loading " + filepath + "...")
    wb = load_workbook(filepath, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.values)
        if len(rows) < 2:
            continue
        df = pd.DataFrame(rows[1:], columns=rows[0])
        sheets[name] = df
        print("  Sheet '" + name + "': " + str(len(df)) + " rows")
    return sheets

# ── CLEAN ─────────────────────────────────────────────────────────────────────

NUMERIC_COLS = [
    "Industry SP", "Betfair SP", "Pred ISP",
    "% SP Drop", "Runs last 18 mo", "Days Since LTO",
    "Wins L5", "Avg % SP Drop L5", "Avg % SP Drop last 18 mo",
    "LTO % SP Drop", "LTO2 % SP Drop", "LTO3 % SP Drop",
    "LTO4 % SP Drop", "LTO5 % SP Drop",
    "LTO IPL", "LTO2 IPL", "LTO3 IPL", "LTO4 IPL", "LTO5 IPL",
    "Crs Wins", "Dist Wins", "Class Wins", "Going Wins",
    "Dist (F)", "WGT (Lbs)", "WGT diff since LTO",
    "Cla diff since LTO", "OR diff since LTO",
    "PRB To Date", "DOB %", "DOB P/L £1",
    "10 B2L To Date", "25 B2L To Date", "50 B2L To Date",
    "OR", "Runners", "Age", "Pace", "Stall",
    "LTO Pos", "Prev Races",
]

# Features that are available BEFORE the race runs (safe to use for prediction)
PRE_RACE_FEATURES = [
    # Market prediction signal
    "Pred ISP", "Industry SP",
    "pred_vs_industry",           # how much model disagrees with market
    "market_rank_in_race",        # relative market position in field

    # LTO market history (last 5 races market support)
    "LTO % SP Drop", "LTO2 % SP Drop", "LTO3 % SP Drop",
    "LTO4 % SP Drop", "LTO5 % SP Drop",
    "avg_lto_sp_drop",            # average SP drop across last 5
    "lto_sp_drop_trend",          # is support improving or declining?
    "consistent_market_support",  # was backed in 3+ of last 5?

    # In-play price history (how low did it trade in last runs)
    "LTO IPL", "LTO2 IPL", "LTO3 IPL", "LTO4 IPL", "LTO5 IPL",
    "avg_lto_ipl",
    "best_lto_ipl",

    # Form
    "Wins L5",
    "Avg % SP Drop L5",
    "Avg % SP Drop last 18 mo",
    "win_rate_l5",
    "LTO Pos",

    # Horse profile
    "PRB To Date",                # historical Percentage Run Better
    "DOB %",                      # Day of Birth % (breeding metric)
    "DOB P/L £1",
    "10 B2L To Date",             # historical beaten by leader stats
    "Runs last 18 mo",
    "Days Since LTO",
    "fresh",                      # 14-42 days since last run
    "stale",                      # 90+ days since last run
    "Age",
    "OR",                         # official rating
    "WGT (Lbs)",
    "WGT diff since LTO",
    "Dist (F)",

    # Affinity
    "Crs Wins", "Dist Wins", "Going Wins",
    "course_winner_flag", "distance_winner_flag",
    "Cla diff since LTO",         # class change
    "OR diff since LTO",          # rating change
    "class_dropped",
    "or_improved",

    # Race context
    "Runners",
    "Pace",
    "Stall",
]


def clean_and_engineer(df):
    df = df.copy()

    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "Place" in df.columns:
        df["Place_num"] = pd.to_numeric(df["Place"], errors="coerce")

    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # ── Feature engineering ───────────────────────────────────────────────────

    if "Pred ISP" in df.columns and "Industry SP" in df.columns:
        df["pred_vs_industry"] = df["Pred ISP"] - df["Industry SP"]

    lto_sp_cols = [c for c in ["LTO % SP Drop","LTO2 % SP Drop","LTO3 % SP Drop",
                                "LTO4 % SP Drop","LTO5 % SP Drop"] if c in df.columns]
    if lto_sp_cols:
        df["avg_lto_sp_drop"] = df[lto_sp_cols].mean(axis=1)
        # Trend: is latest SP drop higher than earlier ones?
        if "LTO % SP Drop" in df.columns and "LTO3 % SP Drop" in df.columns:
            df["lto_sp_drop_trend"] = df["LTO % SP Drop"] - df["LTO3 % SP Drop"]
        # Consistent support: backed (SP drop > 20) in majority of last 5
        support_flags = [(df[c] > 20).astype(int) for c in lto_sp_cols if c in df.columns]
        if support_flags:
            df["consistent_market_support"] = sum(support_flags) >= 3

    lto_ipl_cols = [c for c in ["LTO IPL","LTO2 IPL","LTO3 IPL","LTO4 IPL","LTO5 IPL"] if c in df.columns]
    if lto_ipl_cols:
        df["avg_lto_ipl"] = df[lto_ipl_cols].mean(axis=1)
        df["best_lto_ipl"] = df[lto_ipl_cols].min(axis=1)  # lowest = traded closest to SP

    if "Wins L5" in df.columns:
        df["win_rate_l5"] = df["Wins L5"] / 5.0

    if "Days Since LTO" in df.columns:
        df["fresh"] = df["Days Since LTO"].between(14, 42).astype(int)
        df["stale"] = (df["Days Since LTO"] > 90).astype(int)

    if "Course Winner" in df.columns:
        df["course_winner_flag"] = (df["Course Winner"].astype(str).str.upper() == "YES").astype(int)
    if "Distance Winner" in df.columns:
        df["distance_winner_flag"] = (df["Distance Winner"].astype(str).str.upper() == "YES").astype(int)

    if "Cla diff since LTO" in df.columns:
        df["class_dropped"] = (df["Cla diff since LTO"] > 0).astype(int)
    if "OR diff since LTO" in df.columns:
        df["or_improved"] = (df["OR diff since LTO"] > 0).astype(int)

    # Market rank within race (lower rank = market fav)
    if "Industry SP" in df.columns and "Todays Race" in df.columns:
        df["market_rank_in_race"] = df.groupby("Todays Race")["Industry SP"].rank(method="min")

    return df

# ── SPLIT ─────────────────────────────────────────────────────────────────────

def split_hist_today(df):
    df = clean_and_engineer(df)
    hist  = df[df["Place_num"].notna()].copy()
    today = df[df["Place_num"].isna()].copy()
    hist["won"] = (hist["Place_num"] == 1).astype(int)
    print("  Historical rows : " + str(len(hist)) + "  |  Winners: " + str(int(hist["won"].sum())))
    print("  Today's runners : " + str(len(today)))
    return hist, today

# ── TRAIN ─────────────────────────────────────────────────────────────────────

def train(hist):
    available = [c for c in PRE_RACE_FEATURES if c in hist.columns]
    X = hist[available].copy()
    y = hist["won"].copy()

    medians = X.median()
    X = X.fillna(medians)

    # Remove rows with too many missing values
    valid = X.notna().sum(axis=1) >= int(len(available) * 0.3)
    X, y = X[valid], y[valid]

    print("  Features used   : " + str(len(available)))
    print("  Training rows   : " + str(len(X)))
    print("  Win rate        : " + str(round(y.mean() * 100, 1)) + "%")

    scaler  = StandardScaler()
    Xs      = scaler.fit_transform(X)

    # Gradient Boosting with conservative settings to avoid overfitting
    model = GradientBoostingClassifier(
        n_estimators=300,
        max_depth=3,
        learning_rate=0.03,
        subsample=0.7,
        min_samples_leaf=30,
        max_features=0.6,
        random_state=42,
    )

    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
    cv_scores = cross_val_score(model, Xs, y, cv=cv, scoring="roc_auc")
    print("  Cross-val AUC   : " + str(round(cv_scores.mean(), 3)) +
          "  (+/- " + str(round(cv_scores.std(), 3)) + ")")

    if cv_scores.mean() < 0.52:
        print("  WARNING: Model barely better than random. More data needed.")

    model.fit(Xs, y)

    imp = pd.Series(model.feature_importances_, index=available).nlargest(12)
    print("\n  Most predictive features:")
    for f, v in imp.items():
        bar = "█" * max(1, int(v * 300))
        print("    {:<35} {:.3f}  {}".format(f, v, bar[:40]))

    return model, scaler, available, medians

# ── SCORE ─────────────────────────────────────────────────────────────────────

def score(today, model, scaler, features, medians):
    if len(today) == 0:
        return today
    X = today[features].copy().fillna(medians)
    today = today.copy()
    today["win_prob"] = model.predict_proba(scaler.transform(X))[:, 1]
    return today

# ── RECOMMENDATIONS ───────────────────────────────────────────────────────────

def recommend(scored):
    df = scored.copy()

    # Use Industry SP as our odds source (Betfair SP not available pre-race)
    df["odds"] = pd.to_numeric(df["Industry SP"], errors="coerce")
    df["market_prob"] = 1.0 / df["odds"].clip(lower=1.01)

    df["back_edge"] = df["win_prob"] - df["market_prob"]
    df["lay_edge"]  = df["market_prob"] - df["win_prob"]

    # Expected Value
    df["back_ev"] = (df["win_prob"] * (df["odds"] - 1)) - (1 - df["win_prob"])
    df["lay_ev"]  = ((1 - df["win_prob"]) * (1 - BETFAIR_COMM)) - (df["win_prob"] * (df["odds"] - 1))

    # Fractional Kelly stake (50% Kelly for safety, scaled to 100-unit bank)
    df["back_kelly"]  = ((df["win_prob"] * df["odds"] - 1) / (df["odds"] - 1)).clip(0, 0.25)
    df["back_stake"]  = (df["back_kelly"] * 0.5 * 100).round(2)
    df["lay_kelly"]   = (df["lay_edge"] / (df["odds"] - 1)).clip(0, 0.10)
    df["lay_liab"]    = (df["lay_kelly"] * 0.5 * 100).round(2)

    back = df[
        (df["back_edge"] >= MIN_EDGE) &
        (df["odds"] >= MIN_BACK_ODDS) &
        (df["odds"] <= MAX_BACK_ODDS) &
        (df["back_ev"] > 0)
    ].sort_values("back_ev", ascending=False).head(TOP_N_BETS)

    lay = df[
        (df["lay_edge"] >= MIN_EDGE) &
        (df["odds"] >= MIN_LAY_ODDS) &
        (df["odds"] <= MAX_LAY_ODDS) &
        (df["lay_ev"] > 0)
    ].sort_values("lay_ev", ascending=False).head(TOP_N_BETS)

    return back, lay, df

# ── PRINT ─────────────────────────────────────────────────────────────────────

def fmt_pct(v):
    return str(round(v * 100, 1)) + "%"

def fmt_odds(v):
    return str(round(v, 2)) if pd.notna(v) else "?"

def signals(row, bet_type):
    s = []
    if pd.notna(row.get("avg_lto_sp_drop")) and row["avg_lto_sp_drop"] >= 40:
        s.append("Consistently backed (" + str(round(row["avg_lto_sp_drop"], 0)) + "% avg SP drop)")
    if pd.notna(row.get("Wins L5")) and row["Wins L5"] >= 2:
        s.append(str(int(row["Wins L5"])) + " wins from last 5")
    if row.get("course_winner_flag") == 1:
        s.append("Course winner")
    if row.get("distance_winner_flag") == 1:
        s.append("Distance winner")
    if pd.notna(row.get("PRB To Date")) and row["PRB To Date"] >= 0.65:
        s.append("High PRB (" + str(round(row["PRB To Date"], 2)) + ")")
    if pd.notna(row.get("LTO Pos")) and row["LTO Pos"] == 1:
        s.append("Won last time out")
    if row.get("class_dropped") == 1:
        s.append("Dropped in class")
    if bet_type == "lay":
        if pd.notna(row.get("avg_lto_sp_drop")) and row["avg_lto_sp_drop"] <= 0:
            s.append("No market support recently")
        if pd.notna(row.get("Wins L5")) and row["Wins L5"] == 0:
            s.append("0 wins from last 5")
        if pd.notna(row.get("PRB To Date")) and row["PRB To Date"] <= 0.2:
            s.append("Low PRB (" + str(round(row["PRB To Date"], 2)) + ")")
        if row.get("stale") == 1:
            s.append("Long absence")
    return s


def print_report(back, lay, all_today):
    W = 73
    SEP = "-" * W

    races = all_today["Todays Race"].nunique() if "Todays Race" in all_today.columns else "?"
    print("\n" + "=" * W)
    print("  DAILY BETTING RECOMMENDATIONS")
    print("=" * W)
    print("  Races today : " + str(races) + "   Horses scored : " + str(len(all_today)))
    print("=" * W)

    # ── BACK BETS ─────────────────────────────────────────────────────────────
    print("\n  ★  BACK BETS  —  bet ON these horses to WIN")
    print("  Stakes based on 100-unit bank (fractional Kelly, 50%)")
    print("  " + SEP)

    if len(back) == 0:
        print("  No back bets meet minimum edge criteria today.\n")
    else:
        for _, r in back.iterrows():
            print("")
            print("  HORSE  :  " + str(r.get("Horse","")).upper())
            print("  Race   :  " + str(r.get("Todays Race","")))
            if pd.notna(r.get("Racetype")):  print("  Type   :  " + str(r["Racetype"]))
            if pd.notna(r.get("Distance")):  print("  Dist   :  " + str(r["Distance"]))
            if pd.notna(r.get("Going")):     print("  Going  :  " + str(r["Going"]))
            if pd.notna(r.get("Runners")):   print("  Field  :  " + str(int(r["Runners"])) + " runners")
            print("  Odds   :  " + fmt_odds(r["odds"]) + "  (Predicted: " + fmt_odds(r.get("Pred ISP")) + ")")
            print("  Model  :  " + fmt_pct(r["win_prob"]) + " win chance  vs  market's " + fmt_pct(r["market_prob"]))
            print("  Edge   :  +" + str(round(r["back_edge"]*100,1)) + "% in your favour")
            print("  EV     :  +" + str(round(r["back_ev"],3)) + " per £1 staked")
            print("  STAKE  :  £" + str(r["back_stake"]) + " / 100 units  →  profit if wins: £" + str(round(r["back_stake"]*(r["odds"]-1),2)))
            sigs = signals(r, "back")
            if sigs: print("  Why    :  " + "  |  ".join(sigs))
            print("  " + SEP)

    # ── LAY BETS ──────────────────────────────────────────────────────────────
    print("\n  ✗  LAY BETS  —  bet AGAINST these horses (collect if they LOSE)")
    print("  Liability = max you pay out if horse wins. Profit = what you collect if it loses.")
    print("  " + SEP)

    if len(lay) == 0:
        print("  No lay bets meet minimum edge criteria today.\n")
    else:
        for _, r in lay.iterrows():
            print("")
            print("  HORSE    :  " + str(r.get("Horse","")).upper())
            print("  Race     :  " + str(r.get("Todays Race","")))
            if pd.notna(r.get("Racetype")): print("  Type     :  " + str(r["Racetype"]))
            if pd.notna(r.get("Runners")):  print("  Field    :  " + str(int(r["Runners"])) + " runners")
            print("  Lay odds :  " + fmt_odds(r["odds"]))
            print("  Model    :  " + fmt_pct(r["win_prob"]) + " win chance  vs  market's " + fmt_pct(r["market_prob"]))
            print("  Edge     :  +" + str(round(r["lay_edge"]*100,1)) + "% — model says horse is OVERPRICED")
            print("  EV       :  +" + str(round(r["lay_ev"],3)) + " per £1 profit target")
            print("  LIABILITY:  £" + str(r["lay_liab"]) + " / 100 units  →  profit if loses: £" + str(round(r["lay_liab"]/(r["odds"]-1),2)))
            sigs = signals(r, "lay")
            if sigs: print("  Why      :  " + "  |  ".join(sigs))
            print("  " + SEP)

    # ── RACE BREAKDOWN ────────────────────────────────────────────────────────
    print("\n  ALL RACES — horses ranked by model win probability")
    print("  " + SEP)

    if "Todays Race" not in all_today.columns:
        print("  (No race grouping available)")
        return

    for race, grp in all_today.groupby("Todays Race"):
        grp = grp.sort_values("win_prob", ascending=False)
        print("\n  " + str(race))
        print("  {:<22} {:>7} {:>7} {:>7} {:>6} {:>8}".format(
            "Horse", "Model%", "ISP", "Mkt%", "Edge", "Signal"))
        print("  " + "-" * 60)
        for _, r in grp.iterrows():
            h      = str(r.get("Horse",""))[:21]
            mp     = str(round(r["win_prob"]*100,1)) + "%"
            isp    = fmt_odds(r.get("odds"))
            mktp   = str(round(r["market_prob"]*100,1)) + "%" if pd.notna(r.get("market_prob")) else "?"
            edge   = r.get("back_edge", 0)
            edgestr = ("+" if edge > 0 else "") + str(round(edge*100,1)) + "%"
            sig    = ""
            if r.get("consistent_market_support") == 1: sig = "MKTBACKED"
            elif pd.notna(r.get("Wins L5")) and r["Wins L5"] >= 2: sig = "IN FORM"
            elif r.get("course_winner_flag") == 1: sig = "CRS WNR"
            print("  {:<22} {:>7} {:>7} {:>7} {:>6} {:>9}".format(h, mp, isp, mktp, edgestr, sig))

    print("\n" + "=" * W)
    print("  IMPORTANT: This is a probabilistic model — not a guarantee.")
    print("  Track results daily and reassess if win rate drifts below 8%.")
    print("  Never bet more than you can afford to lose.")
    print("=" * W + "\n")

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else "Daily.xlsx"

    print("\n" + "=" * 73)
    print("  HORSE RACING PREDICTION MODEL  v2")
    print("=" * 73 + "\n")

    sheets = load_excel(filepath)

    if "Combined" in sheets:
        df = sheets["Combined"]
        print("\nUsing Combined sheet")
    else:
        parts = [v for k, v in sheets.items() if k not in ["Selections","Combined"]]
        if not parts:
            print("No data sheets found.")
            return
        df = pd.concat(parts, ignore_index=True)
        print("Merged course sheets")

    print("\nPreparing data...")
    hist, today = split_hist_today(df)

    if len(hist) < 100:
        print("Not enough historical data (need 100+ rows). Exiting.")
        return
    if len(today) == 0:
        print("No unresolved runners found in file.")
        print("Make sure today's horses have a blank Place column.")
        return

    print("\nTraining model...")
    model, scaler, features, medians = train(hist)

    print("\nScoring today's runners...")
    scored = score(today, model, scaler, features, medians)

    print("\nBuilding recommendations...")
    back, lay, all_today = recommend(scored)
    print("  Back bets : " + str(len(back)))
    print("  Lay bets  : " + str(len(lay)))

    print_report(back, lay, all_today)


if __name__ == "__main__":
    main()
